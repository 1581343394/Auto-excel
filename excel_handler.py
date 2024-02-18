import os
import tkinter as tk
import pandas as pd
from openpyxl import load_workbook
import openpyxl as op
import csv

class ExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        self.file_extension = os.path.splitext(self.file_path)[-1]
        # 对于Excel文件
        if self.file_extension in [".xlsx", ".xls"]:
            engine = 'openpyxl' if self.file_extension == ".xlsx" else 'xlrd'
            self.data = pd.read_excel(file_path, sheet_name=None, engine=engine)
            if self.file_extension == ".xlsx":
                self.wb = op.load_workbook(file_path)
            else:
                raise ValueError("Openpyxl does not support .xls files. Please convert it to .xlsx format.")
        # 对于CSV文件
        elif self.file_extension == ".csv":
            self.data = pd.read_csv(file_path)
            with open(file_path, 'r') as f:
                reader = csv.reader(f)
                self.wb = list(reader)
        else:
            raise ValueError("Unsupported file format. Please use .xlsx, .xls or .csv file.")

    def close(self):
        if self.file_extension in [".xlsx", ".xls"]:
            self.wb.close()
    def get_sheet_names(self):
        return list(self.data.keys())

    def get_sheet_info(self, sheet_name):
        sheet = self.data[sheet_name]
        return sheet.shape



    def get_columns(self, sheet_name):
        # 获取指定sheet表
        sheet = self.wb[sheet_name]
        # 获取第一行（假设列名在第一行）
        header_row = sheet[1]
        # 获取所有列名
        columns = []
        for cell in header_row:
            columns.append(cell.value)
        return columns

    # 拼接函数
    def modify_column(self, sheet_name, column_name, user_input1, user_input2):
        # 复制整个工作簿
        new_workbook = load_workbook(self.file_path)
        # 获取指定的 sheet
        sheet = new_workbook[sheet_name]
        # 查找列名所在的行
        header_row = None
        for row in sheet.iter_rows():
            if column_name in [cell.value for cell in row]:
                header_row = row
                break

        # 如果找不到列名，抛出一个异常
        if header_row is None:
            raise ValueError(f"Column name '{column_name}' not found in sheet")
        # 创建从列名到列索引的映射
        column_index = {v.value: i for i, v in enumerate(header_row, start=1)}
        # 获取指定的列索引
        column_idx = column_index[column_name]
        # 对每一行进行处理
        for row in sheet.iter_rows(min_row=header_row[0].row + 1):
            # 获取当前行中指定的单元格
            cell = row[column_idx - 1]
            cell.value = user_input1 + str(cell.value) + user_input2
        # 分割文件名和扩展名
        file_name, file_ext = os.path.splitext(self.file_path)
        # 在文件名后面添加 "副本"
        new_file_name = file_name + "副本" + file_ext
        # 保存为一个新的 Excel 文件
        new_workbook.save(new_file_name)
        tk.messagebox.showinfo("信息", f"已完成拼接，新文件保存在 {new_file_name}")

    # 拆分列函数
    def split_column(self, sheet_name, column_name):
        # 复制整个工作簿
        new_workbook = load_workbook(self.file_path)
        # 获取指定的 sheet
        sheet = new_workbook[sheet_name]
        # 找到列名所在的行
        header_row = None
        for row in sheet.iter_rows():
            if column_name in [cell.value for cell in row]:
                header_row = row
                break
        # 如果找不到列名，抛出一个异常
        if header_row is None:
            raise ValueError(f"Column name '{column_name}' not found in sheet")
        # 创建从列名到列索引的映射
        column_index = {v.value: i for i, v in enumerate(header_row, start=1)}
        # 获取指定的列索引
        column_idx = column_index[column_name]
        # 创建一个字典来保存每个值对应的工作表
        sheets = {}
        # 遍历工作表的每一行
        for row in sheet.iter_rows(min_row=header_row[0].row + 1):
            # 获取当前行中指定的单元格的值
            value = row[column_idx - 1].value
            # 如果这个值对应的工作表还没有被创建，就创建一个新的工作表
            if value not in sheets:
                sheets[value] = self.wb.create_sheet(title=value)
            # 将当前行复制到对应的工作表中
            sheets[value].append([cell.value for cell in row])
        # 定义新工作簿的名称和路径
        new_file_path = os.path.join(os.path.dirname(self.file_path), 'split_' + os.path.basename(self.file_path))
        # 保存工作簿
        self.wb.save(new_file_path)
        # 显示一个消息框来通知用户
        tk.messagebox.showinfo("信息", f"已完成拆分，新文件保存在 {new_file_path}")

    def map_column(self, sheet_name1, column_name1, sheet_name2, column_name2, column_name3):
        # 判断选中的列是否为空或者列名重复
        if not column_name1 or not column_name2 or not column_name3 or column_name2 == column_name3:
            tk.messagebox.showinfo("提示", "选择列错误")
            return

        # 读取sheet1和sheet2的数据
        df1 = self.data[sheet_name1]
        df2 = self.data[sheet_name2]

        # 将df2的column2设置为索引
        df2.set_index(column_name2, inplace=True)

        # 将df2的column3映射到df1的column1
        df1[column_name2+'_mapped'] = df1[column_name1].map(df2[column_name3])

        # 分割文件名和扩展名
        file_name, file_ext = os.path.splitext(self.file_path)

        # 在文件名后面添加 "副本"
        new_file_name = file_name + "副本" + file_ext

        # 保存到新的Excel文件
        with pd.ExcelWriter(new_file_name, engine='openpyxl') as writer:
            df1.to_excel(writer, sheet_name=sheet_name1,index=False)

            # 循环保存其他sheet表
            for name, df in self.data.items():
                if name != sheet_name1:
                    df.to_excel(writer, sheet_name=name)

        tk.messagebox.showinfo("提示", "映射成功，新文件保存在 " + new_file_name)
    # 自动编号
    def bh_excel_column(self, sheet_name, selected_column):
        # 获取指定的工作表
        sheet = self.wb[sheet_name]
        # 获取选中列的列号
        column_index = None
        for cell in sheet[1]:
            if cell.value == selected_column:
                column_index = cell.column_letter
                break

        if column_index is None:
            tk.messagebox.showerror("Error", "Column not found")
            return

        # 对选中的列进行编号
        counter = 1
        merged_cells = sheet.merged_cells.ranges
        processed_cells = set()
        for row in range(2, sheet.max_row + 1):
            cell = sheet[column_index + str(row)]
            if cell.coordinate not in processed_cells:
                for merged in merged_cells:
                    if cell.coordinate in merged:
                        sheet[op.utils.get_column_letter(merged.min_col) + str(merged.min_row)].value = counter
                        for row in range(merged.min_row, merged.max_row + 1):
                            for col in range(merged.min_col, merged.max_col + 1):
                                processed_cells.add(op.utils.get_column_letter(col) + str(row))
                        break
                else:
                    cell.value = counter
                counter += 1
        # 分割文件名和扩展名
        file_name, file_ext = os.path.splitext(self.file_path)
        # 在文件名后面添加 "副本"
        new_file_name = file_name + "副本" + file_ext
        # 保存到新的Excel文件
        self.wb.save(new_file_name)
        tk.messagebox.showinfo("提示", "编号成功，新文件保存在 " + new_file_name)
    #获取excel信息
    def generate_excel_description(self,file_path):
        excel_path=file_path
        # 加载Excel文件
        xls = pd.ExcelFile(excel_path)
        # 获取所有sheet的名字
        sheet_names = xls.sheet_names
        excel_description = []
        # 遍历所有的sheet
        for sheet_name in sheet_names:
            # 读取当前sheet
            df = pd.read_excel(xls, sheet_name=sheet_name)

            # 构建当前sheet的描述
            sheet_description = f"Sheet '{sheet_name}' has {df.shape[1]} columns: {', '.join(df.columns.tolist())}."
            excel_description.append(sheet_description)
        # 返回Excel文件的描述
        return "Excel file contains " + str(len(sheet_names)) + " sheets: " + "; ".join(excel_description)

